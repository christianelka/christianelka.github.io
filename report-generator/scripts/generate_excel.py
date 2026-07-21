import json, sys, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

# HF  = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
# HF  = PatternFill(start_color="D1EAF0", end_color="D1EAF0", fill_type="solid")
HF  = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
HFN = Font(bold=True)
TB  = Border(left=Side(style='thin'), right=Side(style='thin'),
             top=Side(style='thin'),  bottom=Side(style='thin'))
BF  = Font(bold=True)
RA  = Alignment(horizontal='right')
LA  = Alignment(horizontal='left')

def num(v):
    try: return int(v)
    except: return v

def sanitize_sheet_name(name):
    name = str(name)[:31]
    for ch in [':', '\\', '/', '?', '*', '[', ']']:
        name = name.replace(ch, '-')
    return name

def write_detail_sheet(wb, base_name, rows, description):
    if not rows:
        return None
    safe_name = sanitize_sheet_name(base_name)
    existing = [s.title for s in wb.worksheets]
    if safe_name in existing:
        safe_name = f"{safe_name}_{len(existing)}"
    ws = wb.create_sheet(title=safe_name)
    ws['A1'] = description
    ws['A1'].font = Font(bold=True, size=11)
    if rows:
        headers = list(rows[0].keys())
        for ci, h in enumerate(headers, 1):
            sc(ws, 2, ci, h, font=HFN, fill=HF, border=TB)
        for ri, rd in enumerate(rows, 3):
            for ci, h in enumerate(headers, 1):
                val = rd.get(h, '')
                sc(ws, ri, ci, val if val not in (0, None, '') else '', border=TB)
    auto_fit(ws)
    return safe_name

def find_raw_key(rows, target):
    if not rows:
        return None
    t = target.lower()
    for k in rows[0].keys():
        if t in k.lower():
            return k
    return None

def hour_label_from_value(date_str):
    if date_str is None or date_str == '':
        return None
    try:
        if isinstance(date_str, datetime):
            dt = date_str
        elif isinstance(date_str, (int, float)):
            from datetime import timedelta
            dt = datetime(1899, 12, 30) + timedelta(days=float(date_str))
        else:
            s = str(date_str).strip()
            dt = None
            for fmt in (
                '%m/%d/%Y %I:%M:%S %p', '%m/%d/%Y %H:%M:%S',
                '%m/%d/%Y %I:%M %p', '%m/%d/%Y %H:%M',
                '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S',
            ):
                try:
                    dt = datetime.strptime(s, fmt)
                    break
                except ValueError:
                    continue
            if dt is None:
                try:
                    from datetime import timedelta
                    dt = datetime(1899, 12, 30) + timedelta(days=float(s))
                except (ValueError, TypeError):
                    return None
        h = dt.hour
        ampm = 'PM' if h >= 12 else 'AM'
        h12 = h % 12
        if h12 == 0:
            h12 = 12
        return f'{h12} {ampm}'
    except Exception:
        return None

def match_field_value(rv, expected):
    exp = str(expected).strip()
    if exp in ('#N/A', '(blank)'):
        if rv is None or str(rv).strip() in ('', 'None', 'null', '#N/A', '(blank)'):
            return True
        return str(rv).strip() == exp
    if rv is None:
        return False
    return str(rv).strip() == exp

def filter_raw(rows, hour_match=None, **filters):
    if not rows:
        return []
    out = []
    for row in rows:
        ok = True
        for field, value in filters.items():
            key = field if field in row else find_raw_key(rows, field)
            if key is None:
                ok = False
                break
            if not match_field_value(row.get(key), value):
                ok = False
                break
        if ok and hour_match:
            sk = find_raw_key([row], 'submit date') or find_raw_key(rows, 'submit date')
            if sk is None:
                ok = False
            else:
                hl = hour_label_from_value(row.get(sk))
                if hl != hour_match:
                    ok = False
        if ok:
            out.append(row)
    return out

def link_cell(cell, sheet_name):
    if not sheet_name:
        return
    cell.hyperlink = f"#'{sheet_name}'!A1"
    prev = cell.font
    cell.font = Font(
        name=getattr(prev, 'name', None),
        size=getattr(prev, 'size', None),
        bold=getattr(prev, 'bold', False),
        italic=getattr(prev, 'italic', False),
        color="000000",
        underline=None,
    )

def sc(ws, r, c, v, font=None, fill=None, border=None, align=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if border: cell.border = border
    if align:  cell.alignment = align
    return cell

def auto_fit(ws):
    for col in ws.columns:
        ml, cl = 0, None
        for cell in col:
            try:
                if hasattr(cell, 'column_letter'): cl = cell.column_letter
                if cell.value: ml = max(ml, len(str(cell.value)))
            except: pass
        if cl: ws.column_dimensions[cl].width = min(ml + 2, 50)

def write_raw_data(ws, rows):
    if not rows: return
    headers = list(rows[0].keys())
    for ci, h in enumerate(headers, 1):
        sc(ws, 1, ci, h, font=HFN, fill=HF, border=TB)
    for ri, rd in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            sc(ws, ri, ci, rd.get(h, ''), border=TB)
    auto_fit(ws)

def split_pivot_keys(records):
    """
    pivotCount returns e.g. {"Ticket Hashtag": "BadTicket", "25384509": 2, "_total": 3}
    The row_key is the one whose VALUE is always a string label (not a number).
    Agent cols are the numeric NIK keys.
    We identify row_key by checking which key has *string* values across ALL records.
    ponytail: collects keys from ALL records, not just records[0], to avoid
    dropping agent columns that only appear in later rows.
    """
    if not records:
        return None, []
    all_keys = set()
    for rd in records:
        all_keys.update(k for k in rd.keys() if k != '_total')
    all_keys = sorted(all_keys)
    row_key = None
    for k in all_keys:
        if all(isinstance(rd.get(k), str) for rd in records if rd.get(k) is not None):
            row_key = k
            break
    if row_key is None:
        row_key = list(all_keys)[-1]
    agent_cols = [k for k in all_keys if k != row_key]
    return row_key, agent_cols

def write_pivot_section(ws, records, title, start_row, start_col, wb=None, raw_rows=None, user_field=None):
    row_key, agent_cols = split_pivot_keys(records)
    records = sorted(records, key=lambda rd: str(rd.get(row_key, '')).lower())
    raw_rows = raw_rows or []
    user_field = user_field or ''

    r = start_row
    sc(ws, r, start_col, title, font=Font(bold=True, size=12))
    r += 1
    sc(ws, r, start_col, "Count of Incident ID*+", font=HFN, fill=HF, border=TB)
    sc(ws, r, start_col + 1, "Column Labels",       font=HFN, fill=HF, border=TB)
    r += 1

    sc(ws, r, start_col, "Row Labels", font=HFN, fill=HF, border=TB)
    c = start_col + 1
    for agent in agent_cols:
        sc(ws, r, c, num(agent), font=HFN, fill=HF, border=TB, align=RA)
        c += 1
    sc(ws, r, c, "Grand Total", font=HFN, fill=HF, border=TB)
    rightmost = c

    for empty_c in range(start_col + 2, rightmost + 1):
        sc(ws, r - 1, empty_c, '', font=HFN, fill=HF, border=TB)

    for rd in records:
        r += 1
        hashtag = rd.get(row_key, '')
        sc(ws, r, start_col, hashtag, border=TB)
        c = start_col + 1
        for agent in agent_cols:
            val = rd.get(agent, '')
            cell = sc(ws, r, c, val if val != 0 else '', border=TB)
            if wb and val and raw_rows:
                filters = {row_key: hashtag}
                if user_field:
                    filters[user_field] = agent
                detail = filter_raw(raw_rows, **filters)
                if detail:
                    desc = f"Details for Count of Incident ID*+ - {row_key}: {hashtag}"
                    if user_field:
                        desc += f", {user_field}: {agent}"
                    sn = write_detail_sheet(wb, f"D_{title[:6]}_{agent}_{hashtag}"[:31], detail, desc)
                    link_cell(cell, sn)
            c += 1
        total_val = rd.get('_total', '')
        cell = sc(ws, r, c, total_val, border=TB, font=BF)
        if wb and total_val and raw_rows:
            detail = filter_raw(raw_rows, **{row_key: hashtag})
            if detail:
                desc = f"Details for Count of Incident ID*+ - {row_key}: {hashtag}"
                sn = write_detail_sheet(wb, f"D_{title[:6]}_T_{hashtag}"[:31], detail, desc)
                link_cell(cell, sn)

    r += 1
    sc(ws, r, start_col, "Grand Total", font=BF, border=TB)
    c = start_col + 1
    for agent in agent_cols:
        t = sum(rd.get(agent, 0) for rd in records if isinstance(rd.get(agent), (int, float)))
        cell = sc(ws, r, c, t if t != 0 else '', font=BF, border=TB)
        if wb and t and raw_rows and user_field:
            detail = filter_raw(raw_rows, **{user_field: agent})
            if detail:
                desc = f"Details for Count of Incident ID*+ - {user_field}: {agent}"
                sn = write_detail_sheet(wb, f"D_{title[:6]}_GT_{agent}"[:31], detail, desc)
                link_cell(cell, sn)
        c += 1
    sc(ws, r, c, sum(rd.get('_total', 0) for rd in records), font=BF, border=TB)

    return r, rightmost

def write_simple_section(ws, records, title, uk, start_row, start_col, wb=None, raw_rows=None):
    r = start_row
    sc(ws, r, start_col, title, font=Font(bold=True, size=12))
    r += 1
    sc(ws, r, start_col,     "Row Labels",            font=HFN, fill=HF, border=TB)
    sc(ws, r, start_col + 1, "Count of Incident ID*+", font=HFN, fill=HF, border=TB)
    raw_rows = raw_rows or []
    for rd in records:
        r += 1
        user_val = rd.get(uk, '')
        sc(ws, r, start_col,     num(user_val), border=TB, align=LA)
        cell = sc(ws, r, start_col + 1, rd.get('count', ''), border=TB)
        if wb and rd.get('count') and raw_rows:
            detail = filter_raw(raw_rows, **{uk: user_val})
            if detail:
                desc = f"Details for Count of Incident ID*+ - {uk}: {user_val}"
                sn = write_detail_sheet(wb, f"D_{title[:8]}_{user_val}"[:31], detail, desc)
                link_cell(cell, sn)
    r += 1
    sc(ws, r, start_col,     "Grand Total", font=BF, border=TB)
    sc(ws, r, start_col + 1, sum(rd.get('count', 0) for rd in records), font=BF, border=TB)
    return r, start_col + 1

def _ola_sort_key(col, night_shift=False):
    fixed = {'Met': 0, 'Missed': 1, '#N/A': 2, '(blank)': 3}
    if col in fixed:
        return (0, fixed[col])
    s = col.strip().upper()
    if s.endswith('AM') or s.endswith('PM'):
        try:
            h = int(s[:-2].strip()) % 12
            if s.endswith('PM'):
                h += 12
            if night_shift and h >= 12:
                h -= 24
            return (1, h)
        except ValueError:
            pass
    return (2, col)

def write_ola_section(ws, records, ola_date, start_row, start_col, wb=None, raw_rows=None):
    if not records:
        return start_row, start_col
    raw_rows = raw_rows or []
    ola_col_order = ['Met', 'Missed', '#N/A', '(blank)']
    all_ola = set()
    for rd in records:
        for k in rd.keys():
            if k not in ['Row Labels', '_total']:
                all_ola.add(k)
    val_cols = [c for c in ola_col_order if c in all_ola]
    hours = []
    for c in all_ola:
        s = c.strip().upper()
        if s.endswith('AM') or s.endswith('PM'):
            try:
                h = int(s[:-2].strip()) % 12
                if s.endswith('PM'):
                    h += 12
                hours.append(h)
            except ValueError:
                pass
    night_shift = len(hours) > 1 and (max(hours) - min(hours)) > 12
    for c in sorted(all_ola, key=lambda col: _ola_sort_key(col, night_shift)):
        if c not in val_cols: val_cols.append(c)

    r = start_row
    sc(ws, r, start_col, "OLA Response", font=Font(bold=True, size=12))
    r += 1
    sc(ws, r, start_col,     "Count of Incident ID*+", font=HFN, fill=HF, border=TB)
    sc(ws, r, start_col + 1, "Column Labels",          font=HFN, fill=HF, border=TB)
    r += 1
    sc(ws, r, start_col, "Row Labels", font=HFN, fill=HF, border=TB)
    c = start_col + 1
    for cn in val_cols:
        sc(ws, r, c, cn, font=HFN, fill=HF, border=TB); c += 1
    sc(ws, r, c, "Grand Total", font=HFN, fill=HF, border=TB)
    rightmost = c

    for empty_c in range(start_col + 2, rightmost + 1):
        sc(ws, r - 1, empty_c, '', font=HFN, fill=HF, border=TB)

    r += 1
    sc(ws, r, start_col, f"<{ola_date}", font=Font(italic=True), border=TB)
    for rd in records:
        r += 1
        hour_val = rd.get('Row Labels', '')
        sc(ws, r, start_col, hour_val, font=BF, border=TB)
        c = start_col + 1
        for cn in val_cols:
            val = rd.get(cn, '')
            cell = sc(ws, r, c, val if val != 0 else '', border=TB)
            if wb and val and raw_rows:
                detail = filter_raw(raw_rows, hour_match=hour_val, **{'OLA Response': cn})
                if detail:
                    desc = f"Details for Count of Incident ID*+ - Hours (Submit Date): {hour_val}, OLA Response: {cn}"
                    sn = write_detail_sheet(wb, f"D_OLA_{hour_val}_{cn}"[:31], detail, desc)
                    link_cell(cell, sn)
            c += 1
        cell = sc(ws, r, c, rd.get('_total', ''), border=TB, font=BF)
        if wb and rd.get('_total') and raw_rows:
            detail = filter_raw(raw_rows, hour_match=hour_val)
            if detail:
                desc = f"Details for Count of Incident ID*+ - Hours (Submit Date): {hour_val}"
                sn = write_detail_sheet(wb, f"D_OLA_T_{hour_val}"[:31], detail, desc)
                link_cell(cell, sn)

    r += 1
    sc(ws, r, start_col, "Grand Total", font=BF, border=TB)
    c = start_col + 1
    for cn in val_cols:
        t = sum(rd.get(cn, 0) for rd in records if isinstance(rd.get(cn), (int, float)))
        cell = sc(ws, r, c, t if t != 0 else '', font=BF, border=TB)
        if wb and t and raw_rows:
            detail = filter_raw(raw_rows, **{'OLA Response': cn})
            if detail:
                desc = f"Details for Count of Incident ID*+ - OLA Response: {cn}"
                sn = write_detail_sheet(wb, f"D_OLA_GT_{cn}"[:31], detail, desc)
                link_cell(cell, sn)
        c += 1
    sc(ws, r, c, sum(rd.get('_total', 0) for rd in records), font=BF, border=TB)
    return r, rightmost

def write_agent_section(ws, agents, start_row, start_col):
    r = start_row
    sc(ws, r, start_col, "Agent", font=Font(bold=True, size=12))
    r += 1
    sc(ws, r, start_col,     "Domain", font=HFN, fill=HF, border=TB)
    sc(ws, r, start_col + 1, "Nama",   font=HFN, fill=HF, border=TB)
    for a in agents:
        r += 1
        sc(ws, r, start_col,     num(a.get('nik', '')),  border=TB, align=LA)
        sc(ws, r, start_col + 1, a.get('name', ''), border=TB)
    return r, start_col + 1

def write_top5_section(ws, records, start_row, start_col, wb=None, raw_rows=None):
    r = start_row
    sc(ws, r, start_col, "TOP 5", font=Font(bold=True, size=12))
    r += 1
    sc(ws, r, start_col,     "Row Labels",             font=HFN, fill=HF, border=TB)
    sc(ws, r, start_col + 1, "Count of Incident ID*+", font=HFN, fill=HF, border=TB)
    raw_rows = raw_rows or []
    current_parent = None
    for rd in records:
        r += 1
        label = rd.get('category', '')
        is_parent = rd.get('isParent')
        if is_parent:
            current_parent = label
        display = label if is_parent else f"  {label}"
        sc(ws, r, start_col,     display,             border=TB, font=BF if is_parent else None)
        cell = sc(ws, r, start_col + 1, rd.get('count', ''), border=TB)
        if wb and rd.get('count') and raw_rows:
            if is_parent:
                detail = filter_raw(raw_rows, **{'Operational Categorization Tier 1': label})
                desc = f"Details for Count of Incident ID*+ - Operational Categorization Tier 1+: {label}"
            else:
                detail = filter_raw(raw_rows, **{
                    'Operational Categorization Tier 1': current_parent or '',
                    'Operational Categorization Tier 2': label,
                })
                desc = f"Details for Count of Incident ID*+ - Operational Categorization Tier 1+: {current_parent}, Operational Categorization Tier 2: {label}"
            if detail:
                sn = write_detail_sheet(wb, f"D_TOP_{label}"[:31], detail, desc)
                link_cell(cell, sn)
    r += 1
    sc(ws, r, start_col,     "Grand Total", font=BF, border=TB)
    sc(ws, r, start_col + 1, sum(rd.get('count', 0) for rd in records if rd.get('isParent')), font=BF, border=TB)
    return r, start_col + 1

def create_report(data, output_path, report_date, agents=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot Table"

    ws_rpt = wb.create_sheet("Report - Used")
    write_raw_data(ws_rpt, data.get('rawReport', []))

    ws_ola_sheet = wb.create_sheet("OLA Response")
    write_raw_data(ws_ola_sheet, data.get('rawSLA', []))

    LEFT_START_COL = 2
    LEFT_START_ROW = 1

    cur_row = LEFT_START_ROW
    max_right_left = LEFT_START_COL

    raw_report = data.get('rawReport', [])

    cancelled = data.get('cancelled', [])
    if cancelled:
        last_r, last_c = write_pivot_section(
            ws, cancelled, "Cancelled", cur_row, LEFT_START_COL,
            wb=wb, raw_rows=raw_report, user_field='Status History.Cancelled.USER')
        cur_row = last_r + 2
        max_right_left = max(max_right_left, last_c)

    resolved = data.get('resolved', [])
    if resolved:
        last_r, last_c = write_pivot_section(
            ws, resolved, "Resolved", cur_row, LEFT_START_COL,
            wb=wb, raw_rows=raw_report, user_field='Status History.Resolved.USER')
        cur_row = last_r + 2
        max_right_left = max(max_right_left, last_c)

    inprogress = data.get('inprogress', [])
    assigned = data.get('assigned', [])
    resolved_row_count = len(data.get('resolved', []))

    if resolved_row_count >= 5 and inprogress and assigned:
        uk_ip = list(inprogress[0].keys())[0]
        uk_as = list(assigned[0].keys())[0]
        assigned_col = LEFT_START_COL + 3

        last_r_ip, last_c_ip = write_simple_section(
            ws, inprogress, "Inprogress", uk_ip, cur_row, LEFT_START_COL,
            wb=wb, raw_rows=raw_report)
        last_r_as, last_c_as = write_simple_section(
            ws, assigned, "Assigned", uk_as, cur_row, assigned_col,
            wb=wb, raw_rows=raw_report)

        cur_row = max(last_r_ip, last_r_as) + 2
        max_right_left = max(max_right_left, last_c_ip, last_c_as)
    else:
        if inprogress:
            uk = list(inprogress[0].keys())[0]
            last_r, last_c = write_simple_section(
                ws, inprogress, "Inprogress", uk, cur_row, LEFT_START_COL,
                wb=wb, raw_rows=raw_report)
            cur_row = last_r + 2
            max_right_left = max(max_right_left, last_c)

        if assigned:
            uk = list(assigned[0].keys())[0]
            last_r, last_c = write_simple_section(
                ws, assigned, "Assigned", uk, cur_row, LEFT_START_COL,
                wb=wb, raw_rows=raw_report)
            cur_row = last_r + 2
            max_right_left = max(max_right_left, last_c)

    MID_START_COL = max_right_left + 2

    mid_row = LEFT_START_ROW
    max_right_mid = MID_START_COL

    ola = data.get('olaResponse', [])
    ola_date = data.get('olaDate', report_date)
    if ola:
        last_r, last_c = write_ola_section(
            ws, ola, ola_date, mid_row, MID_START_COL,
            wb=wb, raw_rows=raw_report)
        mid_row = last_r + 2
        max_right_mid = max(max_right_mid, last_c)

    if agents:
        last_r, last_c = write_agent_section(ws, agents, mid_row, MID_START_COL)
        mid_row = last_r + 2
        max_right_mid = max(max_right_mid, last_c)

    RIGHT_START_COL = max_right_mid + 2

    top5 = data.get('topCategories', [])
    if top5:
        write_top5_section(ws, top5, LEFT_START_ROW, RIGHT_START_COL, wb=wb, raw_rows=raw_report)

    auto_fit(ws)
    import os
    tmp_path = output_path + '.tmp'
    wb.save(tmp_path)
    os.replace(tmp_path, output_path)

    csv_path = output_path.replace('.xlsx', '.csv')
    csv_tmp = csv_path + '.tmp'
    generate_csv(data, csv_tmp, agents, MID_START_COL, RIGHT_START_COL)
    os.replace(csv_tmp, csv_path)
    return output_path

def generate_csv(data, csv_path, agents, mid_col, right_col):
    from collections import defaultdict
    sheet = defaultdict(dict)

    def put(r, c, v):
        sheet[r][c] = v

    LEFT = 2
    cur_row = 1
    max_right_left = LEFT

    def put_pivot(records, title, start_row, sc_):
        nonlocal max_right_left
        row_key, agent_cols = split_pivot_keys(records)
        records = sorted(records, key=lambda rd: str(rd.get(row_key, '')).lower())
        r = start_row
        put(r, sc_, title); r += 1
        put(r, sc_, "Count of Incident ID*+"); put(r, sc_ + 1, "Column Labels"); r += 1
        put(r, sc_, "Row Labels")
        c = sc_ + 1
        for a in agent_cols: put(r, c, num(a)); c += 1
        put(r, c, "Grand Total"); max_right_left = max(max_right_left, c)
        for rd in records:
            r += 1
            put(r, sc_, rd.get(row_key, ''))
            c = sc_ + 1
            for a in agent_cols:
                val = rd.get(a, '')
                put(r, c, val if val != 0 else ''); c += 1
            put(r, c, rd.get('_total', ''))
        r += 1
        put(r, sc_, "Grand Total")
        c = sc_ + 1
        for a in agent_cols:
            t = sum(rd.get(a, 0) for rd in records if isinstance(rd.get(a), (int, float)))
            put(r, c, t if t != 0 else ''); c += 1
        put(r, c, sum(rd.get('_total', 0) for rd in records))
        return r + 2

    cancelled = data.get('cancelled', [])
    if cancelled: cur_row = put_pivot(cancelled, "Cancelled", cur_row, LEFT)
    resolved   = data.get('resolved', [])
    if resolved:  cur_row = put_pivot(resolved,  "Resolved",  cur_row, LEFT)

    def put_simple(records, title, uk, start_row, col=LEFT):
        nonlocal max_right_left
        r = start_row
        put(r, col, title); r += 1
        put(r, col, "Row Labels"); put(r, col + 1, "Count of Incident ID*+")
        max_right_left = max(max_right_left, col + 1)
        for rd in records:
            r += 1; put(r, col, num(rd.get(uk, ''))); put(r, col + 1, rd.get('count', ''))
        r += 1; put(r, col, "Grand Total"); put(r, col + 1, sum(rd.get('count', 0) for rd in records))
        return r + 2

    inprogress = data.get('inprogress', [])
    assigned = data.get('assigned', [])
    resolved_row_count = len(data.get('resolved', []))

    if resolved_row_count >= 5 and inprogress and assigned:
        uk_ip = list(inprogress[0].keys())[0]
        uk_as = list(assigned[0].keys())[0]
        assigned_col = LEFT + 3
        r_ip = put_simple(inprogress, "Inprogress", uk_ip, cur_row, LEFT)
        r_as = put_simple(assigned, "Assigned", uk_as, cur_row, assigned_col)
        cur_row = max(r_ip, r_as)
    else:
        if inprogress:
            uk = list(inprogress[0].keys())[0]
            cur_row = put_simple(inprogress, "Inprogress", uk, cur_row)
        if assigned:
            uk = list(assigned[0].keys())[0]
            cur_row = put_simple(assigned, "Assigned", uk, cur_row)

    MC = mid_col
    mid_row = 1
    max_right_mid = MC

    ola = data.get('olaResponse', [])
    ola_date = data.get('olaDate', '')
    if ola:
        ola_col_order = ['Met', 'Missed', '#N/A', '(blank)']
        all_ola = set()
        for rd in ola:
            for k in rd.keys():
                if k not in ['Row Labels', '_total']: all_ola.add(k)
        val_cols = [c for c in ola_col_order if c in all_ola]
        hours = []
        for c in all_ola:
            s = c.strip().upper()
            if s.endswith('AM') or s.endswith('PM'):
                try:
                    h = int(s[:-2].strip()) % 12
                    if s.endswith('PM'):
                        h += 12
                    hours.append(h)
                except ValueError:
                    pass
        night_shift = len(hours) > 1 and (max(hours) - min(hours)) > 12
        for c in sorted(all_ola, key=lambda col: _ola_sort_key(col, night_shift)):
            if c not in val_cols: val_cols.append(c)

        r = mid_row
        put(r, MC, "OLA Response"); r += 1
        put(r, MC, "Count of Incident ID*+"); put(r, MC + 1, "Column Labels"); r += 1
        put(r, MC, "Row Labels")
        c = MC + 1
        for cn in val_cols: put(r, c, cn); c += 1
        put(r, c, "Grand Total"); max_right_mid = max(max_right_mid, c)
        r += 1; put(r, MC, f"<{ola_date}")
        for rd in ola:
            r += 1; put(r, MC, rd.get('Row Labels', ''))
            c = MC + 1
            for cn in val_cols:
                val = rd.get(cn, '')
                put(r, c, val if val != 0 else ''); c += 1
            put(r, c, rd.get('_total', ''))
        r += 1; put(r, MC, "Grand Total")
        c = MC + 1
        for cn in val_cols:
            t = sum(rd.get(cn, 0) for rd in ola if isinstance(rd.get(cn), (int, float)))
            put(r, c, t if t != 0 else ''); c += 1
        put(r, c, sum(rd.get('_total', 0) for rd in ola))
        mid_row = r + 2

    if agents:
        r = mid_row
        put(r, MC, "Agent"); r += 1
        put(r, MC, "Domain"); put(r, MC + 1, "Nama")
        max_right_mid = max(max_right_mid, MC + 1)
        for a in agents:
            r += 1; put(r, MC, num(a.get('nik', ''))); put(r, MC + 1, a.get('name', ''))
        mid_row = r + 2

    RC = right_col
    top5 = data.get('topCategories', [])
    if top5:
        r = 1
        put(r, RC, "TOP 5"); r += 1
        put(r, RC, "Row Labels"); put(r, RC + 1, "Count of Incident ID*+")
        for rd in top5:
            r += 1
            label = rd.get('category', '')
            if not rd.get('isParent'): label = f"  {label}"
            put(r, RC, label); put(r, RC + 1, rd.get('count', ''))
        r += 1; put(r, RC, "Grand Total")
        put(r, RC + 1, sum(rd.get('count', 0) for rd in top5 if rd.get('isParent')))

    max_r = max(sheet.keys()) if sheet else 1
    max_c = max(c for row in sheet.values() for c in row.keys()) if sheet else 1

    grid = [[''] * (max_c + 1) for _ in range(max_r + 1)]
    for r, cols in sheet.items():
        for c, v in cols.items():
            grid[r][c] = '' if v is None else str(v)

    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in grid[1:]:
            writer.writerow(row[1:])

if __name__ == "__main__":
    params     = json.loads(sys.stdin.read())
    data       = params.get('data', {})
    out_path   = params.get('outputPath', 'report.xlsx')
    rep_date   = params.get('reportDate', datetime.now().strftime('%Y-%m-%d'))
    agents     = params.get('agents', [])
    result     = create_report(data, out_path, rep_date, agents)
    print(json.dumps({"success": True, "filePath": result}))
